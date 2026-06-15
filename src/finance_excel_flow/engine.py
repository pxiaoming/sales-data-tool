from __future__ import annotations

from dataclasses import asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .config import AppConfig, BusinessRuleConfig
from .paths import resolve_bundle_path


def _is_empty(value: Any) -> bool:
    return value is None or value == ""


def _as_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if _is_empty(value):
        return default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, (datetime, date)):
        raise TypeError("Cannot convert date value to number")
    try:
        text = str(value).strip().replace(",", "")
        return Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Cannot convert value to number: {value!r}") from exc


def _as_text(value: Any, default: str = "") -> str:
    if _is_empty(value):
        return default
    return str(value)


def _as_date(value: Any, default: date | None = None) -> date | None:
    if _is_empty(value):
        return default
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date value: {value!r}")


def _build_row_context(row: dict[str, Any]) -> dict[str, Any]:
    def lit(value: Any) -> Any:
        return value

    def col(name: str, default: Any = None) -> Any:
        value = row.get(name, default)
        return default if _is_empty(value) else value

    def text(name: str, default: str = "") -> str:
        return _as_text(row.get(name), default)

    def num(name: str, default: Any = 0) -> Decimal:
        return _as_decimal(row.get(name), Decimal(str(default)))

    def date_value(name: str, default: date | None = None) -> date | None:
        return _as_date(row.get(name), default)

    def ifnull(value: Any, default: Any = "") -> Any:
        return default if _is_empty(value) else value

    def contains(name: str, needle: Any) -> bool:
        return str(needle) in text(name)

    def startswith(name: str, prefix: Any) -> bool:
        return text(name).startswith(str(prefix))

    def endswith(name: str, suffix: Any) -> bool:
        return text(name).endswith(str(suffix))

    return {
        "col": col,
        "text": text,
        "num": num,
        "date": date_value,
        "ifnull": ifnull,
        "contains": contains,
        "startswith": startswith,
        "endswith": endswith,
        "lit": lit,
        "round": round,
        "Decimal": Decimal,
        "date_cls": date,
        "datetime_cls": datetime,
    }


def _build_rule_context(summary: dict[str, Any]) -> dict[str, Any]:
    context = {
        "lit": lambda value: value,
        "round": round,
        "Decimal": Decimal,
        "date_cls": date,
        "datetime_cls": datetime,
    }
    context.update(summary)
    return context


def _field_alias(field_name: str) -> str:
    alias_chars: list[str] = []
    for char in field_name.strip():
        if char.isalnum():
            alias_chars.append(char.lower())
        else:
            alias_chars.append("_")
    alias = "".join(alias_chars).strip("_")
    while "__" in alias:
        alias = alias.replace("__", "_")
    return alias or "field"


def _evaluate(expression: str, context: dict[str, Any]) -> Any:
    safe_globals = {"__builtins__": {}}
    return eval(expression, safe_globals, context)


def _group_rows(rows: list[dict[str, Any]], group_by: list[str]) -> list[tuple[tuple[Any, ...], list[dict[str, Any]]]]:
    if not group_by:
        return [(tuple(), rows)]

    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for row in rows:
        key = tuple(row.get(field) for field in group_by)
        grouped.setdefault(key, []).append(row)

    def sort_key(item: tuple[tuple[Any, ...], list[dict[str, Any]]]) -> tuple[str, ...]:
        key, _ = item
        return tuple("" if value is None else str(value) for value in key)

    return sorted(grouped.items(), key=sort_key)


def _read_sheet_rows(input_path: Path, sheet_name: str | int | None, header_row: int) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(input_path, data_only=True)
    if sheet_name is None:
        worksheet = workbook.active
    elif isinstance(sheet_name, int):
        worksheet = workbook.worksheets[sheet_name]
    else:
        worksheet = workbook[sheet_name]

    headers: list[str] = []
    for cell in worksheet[header_row]:
        value = cell.value
        headers.append("" if value is None else str(value).strip())

    rows: list[dict[str, Any]] = []
    for excel_row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[index]: value for index, value in enumerate(excel_row) if index < len(headers) and headers[index]}
        rows.append(row)

    return headers, rows


def _drop_empty_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cleaned: list[dict[str, Any]] = []
    for row in rows:
        if any(not _is_empty(value) for value in row.values()):
            cleaned.append(row)
    return cleaned


def _rename_columns(rows: list[dict[str, Any]], headers: list[str], mapping: dict[str, str]) -> tuple[list[dict[str, Any]], list[str]]:
    if not mapping:
        return rows, headers

    renamed_headers = [mapping.get(header, header) for header in headers]
    if len(set(renamed_headers)) != len(renamed_headers):
        raise ValueError("Renaming columns produced duplicate header names")
    renamed_rows: list[dict[str, Any]] = []
    for row in rows:
        renamed = {}
        for key, value in row.items():
            renamed[mapping.get(key, key)] = value
        renamed_rows.append(renamed)
    return renamed_rows, renamed_headers


def _validate_required_columns(headers: list[str], required_columns: list[str]) -> None:
    if not required_columns:
        return
    available = set(headers)
    missing = [column for column in required_columns if column not in available]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")


def _apply_filters(rows: list[dict[str, Any]], filters: list[str]) -> list[dict[str, Any]]:
    if not filters:
        return rows

    filtered_rows: list[dict[str, Any]] = []
    for row in rows:
        keep = True
        for expression in filters:
            result = _evaluate(expression, _build_row_context(row))
            if not bool(result):
                keep = False
                break
        if keep:
            filtered_rows.append(row)
    return filtered_rows


def _apply_calculations(rows: list[dict[str, Any]], calculations: dict[str, str]) -> list[dict[str, Any]]:
    if not calculations:
        return rows

    calculated_rows: list[dict[str, Any]] = []
    for row in rows:
        new_row = dict(row)
        for column_name, expression in calculations.items():
            new_row[column_name] = _evaluate(expression, _build_row_context(new_row))
        calculated_rows.append(new_row)
    return calculated_rows


def _select_columns(rows: list[dict[str, Any]], headers: list[str], select_columns: list[str]) -> tuple[list[dict[str, Any]], list[str]]:
    if not select_columns:
        return rows, headers

    if len(set(select_columns)) != len(select_columns):
        raise ValueError("select_columns contains duplicates")

    selected_rows: list[dict[str, Any]] = []
    for row in rows:
        selected_rows.append({column: row.get(column) for column in select_columns})
    return selected_rows, select_columns


def _sort_rows(rows: list[dict[str, Any]], sort_by: list[str]) -> list[dict[str, Any]]:
    if not sort_by:
        return rows

    def sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
        normalized: list[Any] = []
        for column in sort_by:
            value = row.get(column)
            if value is None:
                normalized.append((1, 0, ""))
            elif isinstance(value, (int, float, Decimal)):
                normalized.append((0, 0, float(value)))
            elif isinstance(value, (datetime, date)):
                normalized.append((0, 1, value.isoformat()))
            else:
                normalized.append((0, 2, str(value)))
        return tuple(normalized)

    return sorted(rows, key=sort_key)


def _render_output_filename(template: str, input_path: Path) -> str:
    now = datetime.now()
    context = {
        "source_stem": input_path.stem,
        "source_name": input_path.name,
        "source_suffix": input_path.suffix,
        "date": now.date().isoformat(),
        "datetime": now.strftime("%Y-%m-%d %H:%M:%S"),
        "yyyymmdd": now.strftime("%Y%m%d"),
        "yyyymm": now.strftime("%Y%m"),
        "timestamp": now.strftime("%Y%m%d_%H%M%S"),
    }
    return template.format(**context)


def _extract_source_yyyymm(input_path: Path) -> str:
    matches = re.findall(r"(20\d{4})", input_path.stem)
    if matches:
        return matches[-1]
    return datetime.now().strftime("%Y%m")


def _sum_field(rows: list[dict[str, Any]], field_name: str) -> Decimal:
    total = Decimal("0")
    for row in rows:
        total += _as_decimal(row.get(field_name), Decimal("0"))
    return total


def _evaluate_business_rule(rule: BusinessRuleConfig, rows: list[dict[str, Any]], source_path: Path) -> dict[str, Any]:
    matched_rows = _apply_filters(rows, rule.source_filters)
    due_amount_total = _sum_field(matched_rows, rule.sum_field)
    percentage = Decimal(str(rule.percentage))
    accrual_amount = due_amount_total * percentage
    source_yyyymm = _extract_source_yyyymm(source_path)
    run_yyyymm = datetime.now().strftime("%Y%m")
    summary = {
        "rule_name": rule.name,
        "matched_count": len(matched_rows),
        "source_yyyymm": source_yyyymm,
        "run_yyyymm": run_yyyymm,
        "due_amount_total": due_amount_total,
        "percentage": percentage,
        "accrual_amount": accrual_amount,
    }
    summary["name"] = rule.name_template.format(**summary)
    return summary


def _resolve_group_percentage(rule: BusinessRuleConfig, group_key: tuple[Any, ...]) -> Decimal:
    if not rule.group_percentages:
        return Decimal(str(rule.percentage))

    for value in group_key:
        key = str(value)
        if key in rule.group_percentages:
            return Decimal(str(rule.group_percentages[key]))

    return Decimal(str(rule.percentage))


def _evaluate_grouped_business_rule(
    rule: BusinessRuleConfig,
    rows: list[dict[str, Any]],
    source_path: Path,
) -> tuple[list[tuple[dict[str, Any], int | None, int | None]], dict[str, Any] | None]:
    matched_rows = _apply_filters(rows, rule.source_filters)
    grouped_rows = _group_rows(matched_rows, rule.group_by)
    if rule.group_row_pairs and len(grouped_rows) > len(rule.group_row_pairs):
        raise ValueError(
            f"Rule {rule.name} has {len(grouped_rows)} groups but only {len(rule.group_row_pairs)} output row pairs"
        )

    outputs: list[tuple[dict[str, Any], int | None, int | None]] = []
    source_yyyymm = _extract_source_yyyymm(source_path)
    run_yyyymm = datetime.now().strftime("%Y%m")
    total_due_amount = Decimal("0")
    total_accrual_amount = Decimal("0")

    for index, (group_key, group_rows) in enumerate(grouped_rows):
        row_num: int | None = None
        credit_row_num: int | None = None
        if rule.group_row_pairs and index < len(rule.group_row_pairs):
            row_num, credit_row_num = rule.group_row_pairs[index]

        due_amount_total = _sum_field(group_rows, rule.sum_field)
        percentage = _resolve_group_percentage(rule, group_key)
        accrual_amount = due_amount_total * percentage
        total_due_amount += due_amount_total
        total_accrual_amount += accrual_amount

        summary: dict[str, Any] = {
            "rule_name": rule.name,
            "matched_count": len(group_rows),
            "source_yyyymm": source_yyyymm,
            "run_yyyymm": run_yyyymm,
            "due_amount_total": due_amount_total,
            "percentage": percentage,
            "accrual_amount": accrual_amount,
            "group_index": index + 1,
            "row": row_num,
            "credit_row": credit_row_num,
            "group_key": group_key,
        }

        for field_name, value in zip(rule.group_by, group_key, strict=False):
            summary[field_name] = value
            summary[_field_alias(field_name)] = value

        if len(rule.group_by) == 1:
            summary["group_value"] = group_key[0]
            summary["group_value_alias"] = group_key[0]

        summary["name"] = rule.name_template.format(**summary)
        outputs.append((summary, row_num, credit_row_num))

    aggregate_summary: dict[str, Any] | None = None
    if rule.summary_template_cells:
        aggregate_summary = {
            "rule_name": rule.name,
            "matched_count": len(matched_rows),
            "source_yyyymm": source_yyyymm,
            "run_yyyymm": run_yyyymm,
            "due_amount_total": total_due_amount,
            "total_due_amount_total": total_due_amount,
            "accrual_amount": total_accrual_amount,
            "total_accrual_amount": total_accrual_amount,
            "percentage": Decimal(str(rule.percentage)),
        }
        aggregate_summary["name"] = rule.name_template.format(**aggregate_summary)

    return outputs, aggregate_summary


def _auto_fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _write_output(rows: list[dict[str, Any]], headers: list[str], output_path: Path, sheet_name: str, freeze_header: bool, auto_fit_columns: bool) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    if freeze_header and worksheet.max_row >= 1:
        worksheet.freeze_panes = "A2"

    if headers:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    if auto_fit_columns:
        _auto_fit_columns(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_template_output(
    template_path: Path,
    output_path: Path,
    sheet_name: str | None,
    cell_values: dict[str, Any],
    business_rules: list[BusinessRuleConfig],
    source_rows: list[dict[str, Any]],
    source_path: Path,
    freeze_header: bool,
    auto_fit_columns: bool,
) -> None:
    workbook = load_workbook(template_path)
    worksheet = workbook.active if sheet_name is None else workbook[sheet_name]

    for cell_ref, value in cell_values.items():
        worksheet[cell_ref] = value

    for rule in business_rules:
        if rule.group_by:
            grouped_outputs, aggregate_summary = _evaluate_grouped_business_rule(rule, source_rows, source_path)
            for summary, row_num, credit_row_num in grouped_outputs:
                rule_context = _build_rule_context(summary)
                for cell_ref_template, expression in rule.template_cells.items():
                    cell_ref = cell_ref_template.format(**rule_context)
                    worksheet[cell_ref] = _evaluate(expression, rule_context)
            if aggregate_summary is not None:
                aggregate_context = _build_rule_context(aggregate_summary)
                for cell_ref, expression in rule.summary_template_cells.items():
                    worksheet[cell_ref] = _evaluate(expression, aggregate_context)
        else:
            summary = _evaluate_business_rule(rule, source_rows, source_path)
            rule_context = _build_rule_context(summary)
            for cell_ref, expression in rule.template_cells.items():
                worksheet[cell_ref] = _evaluate(expression, rule_context)

    if freeze_header and worksheet.freeze_panes is None:
        worksheet.freeze_panes = "A2"

    if auto_fit_columns:
        _auto_fit_columns(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def process_file(input_path: str | Path, config: AppConfig, output_dir: str | Path | None = None, output_path: str | Path | None = None) -> Path:
    source_path = Path(input_path)
    headers, source_rows = _read_sheet_rows(source_path, config.input.sheet_name, config.input.header_row)
    rows = list(source_rows)

    if config.transform.drop_empty_rows:
        rows = _drop_empty_rows(rows)

    rows, headers = _rename_columns(rows, headers, config.transform.rename_columns)
    _validate_required_columns(headers, config.transform.required_columns)
    rows = _apply_filters(rows, config.transform.filters)
    rows = _apply_calculations(rows, config.transform.calculated_columns)
    headers = list(headers) + [column for column in config.transform.calculated_columns if column not in headers]
    rows = _sort_rows(rows, config.transform.sort_by)
    rows, headers = _select_columns(rows, headers, config.transform.select_columns)

    if output_path is not None:
        target_path = Path(output_path)
    else:
        if output_dir is None:
            output_dir = source_path.parent / "output"
        output_dir = Path(output_dir)
        filename = _render_output_filename(config.output.filename_template, source_path)
        target_path = output_dir / filename

    if config.output.template_path:
        _write_template_output(
            template_path=resolve_bundle_path(config.output.template_path),
            output_path=target_path,
            sheet_name=config.output.sheet_name,
            cell_values=config.output.cell_values,
            business_rules=config.business_rules,
            source_rows=source_rows,
            source_path=source_path,
            freeze_header=config.output.freeze_header,
            auto_fit_columns=config.output.auto_fit_columns,
        )
    else:
        _write_output(
            rows=rows,
            headers=headers,
            output_path=target_path,
            sheet_name=config.output.sheet_name,
            freeze_header=config.output.freeze_header,
            auto_fit_columns=config.output.auto_fit_columns,
        )
    return target_path


def describe_config(config: AppConfig) -> dict[str, Any]:
    return asdict(config)
