from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from openpyxl import Workbook, load_workbook

from finance_excel_flow.config import load_config
from finance_excel_flow.engine import _render_output_filename, process_file


class ConfigTestCase(unittest.TestCase):
    def test_load_config(self) -> None:
        yaml_text = """
input:
  sheet_name: 1
  header_row: 2
output:
  sheet_name: "输出"
  filename_template: "{source_stem}_{yyyymmdd}.xlsx"
transform:
  rename_columns:
    A列: A
  required_columns: [A]
  filters:
    - 'text("A") != ""'
  calculated_columns:
    B: 'num("A") + 1'
  select_columns: [A, B]
  sort_by: [A]
"""
        with tempfile.TemporaryDirectory() as tmpdir:
            config_path = Path(tmpdir) / "config.yml"
            config_path.write_text(yaml_text, encoding="utf-8")
            config = load_config(config_path)

        self.assertEqual(config.input.sheet_name, 1)
        self.assertEqual(config.input.header_row, 2)
        self.assertEqual(config.output.sheet_name, "输出")
        self.assertEqual(config.transform.rename_columns["A列"], "A")
        self.assertEqual(config.transform.required_columns, ["A"])
        self.assertEqual(config.transform.select_columns, ["A", "B"])

    def test_render_output_filename(self) -> None:
        result = _render_output_filename(
            "{source_stem}_{yyyymmdd}.xlsx",
            Path("/tmp/example.xlsx"),
        )
        self.assertTrue(result.startswith("example_"))
        self.assertTrue(result.endswith(".xlsx"))

    def test_template_output_mode(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        template_path = repo_root / "templates" / "ka_promo_and_expense_template.xlsx"

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            input_path = tmpdir_path / "Sales report_NA_202605.xlsx"
            output_path = tmpdir_path / "output.xlsx"
            config_path = tmpdir_path / "config.yml"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Detail"
            worksheet.append(["Customer number", "Due Amount", "Brand"])
            worksheet.append(["0376", 20, "000030"])
            worksheet.append(["0376", 30, "000030"])
            worksheet.append(["0376", 20, "000030"])
            worksheet.append(["B11946", 20, "000030"])
            worksheet.append(["B11946", 20, "000030"])
            worksheet.append(["0314", 20, "000030"])
            worksheet.append(["0314", 20, "000030"])
            worksheet.append(["0314", 20, "000050"])
            worksheet.append(["0314", 20, "000050"])
            worksheet.append(["0314", 20, "000051"])
            worksheet.append(["0314", 20, "000051"])
            worksheet.append(["0314", 20, "000070"])
            worksheet.append(["0314", 20, "000070"])
            worksheet.append(["0314", 20, "000080"])
            worksheet.append(["0314", 20, "000080"])
            worksheet.append(["0178", 20, "000030"])
            worksheet.append(["0178", 20, "000050"])
            worksheet.append(["0178", 20, "000051"])
            worksheet.append(["0178", 20, "000070"])
            worksheet.append(["0178", 20, "000080"])
            worksheet.append(["0178", 20, "ODM002"])
            workbook.save(input_path)

            config_path.write_text(
                f"""
input:
  sheet_name: Detail
  header_row: 1
output:
  sheet_name: Sheet1
  template_path: {template_path.as_posix()}
business_rules:
  - name: ace_ego_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0376"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000030"'
    sum_field: "Due Amount"
    percentage: 0.11
    name_template: "ACE EGO {{source_yyyymm}} Promo Accrual"
    template_cells:
      E5: 'lit("0376")'
      F5: 'lit("000030")'
      H5: 'round(accrual_amount, 2)'
      J5: 'name'
      L5: 'round(due_amount_total, 2)'
      M5: 'percentage'
      E6: 'lit("0376")'
      F6: 'lit("000030")'
      I6: 'round(accrual_amount, 2)'
      J6: 'name'
      L6: 'round(due_amount_total, 2)'
      M6: 'percentage'
  - name: aceego_1pct_marketing_allowance_accrual_1
    source_filters:
      - 'text("Customer number") == "0376"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000030"'
    sum_field: "Due Amount"
    percentage: 0.01
    name_template: "ACEEGO {{source_yyyymm}} 1% Marketing Allowance Accrual"
    template_cells:
      E14: 'lit("0376")'
      F14: 'lit("000030")'
      H11: 'round(accrual_amount, 2)'
      I12: 'round(accrual_amount, 2)'
      H13: 'round(accrual_amount, 2)'
      I14: 'round(accrual_amount, 2)'
      J11: 'name'
      J12: 'name'
      J13: 'name'
      J14: 'name'
      L11: 'round(due_amount_total, 2)'
      L12: 'round(due_amount_total, 2)'
      L13: 'round(due_amount_total, 2)'
      L14: 'round(due_amount_total, 2)'
      M11: 'percentage'
      M12: 'percentage'
      M13: 'percentage'
      M14: 'percentage'
  - name: g10_eg0_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "B11946"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000030"'
    sum_field: "Due Amount"
    percentage: 0.14
    name_template: "G10 EG0 {{source_yyyymm}} Promo Accrual"
    template_cells:
      E19: 'lit("B11946")'
      F19: 'lit("000030")'
      H19: 'round(accrual_amount, 2)'
      I20: 'round(accrual_amount, 2)'
      J19: 'name'
      J20: 'name'
      L19: 'round(due_amount_total, 2)'
      L20: 'round(due_amount_total, 2)'
      M19: 'percentage'
      M20: 'percentage'
  - name: amazon_eg0_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0314"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000030"'
    sum_field: "Due Amount"
    percentage: 0.14
    name_template: "Amazon EG0 {{source_yyyymm}} Promo Accrual"
    template_cells:
      E25: 'lit("0314")'
      F25: 'lit("000030")'
      H25: 'round(accrual_amount, 2)'
      J25: 'name'
      L25: 'round(due_amount_total, 2)'
      M25: 'percentage'
      E30: 'lit("0314")'
      F30: 'lit("000030")'
      I30: 'round(accrual_amount, 2)'
      J30: 'name'
      L30: 'round(due_amount_total, 2)'
      M30: 'percentage'
  - name: amazon_flex_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0314"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000050"'
    sum_field: "Due Amount"
    percentage: 0.145
    name_template: "Amazon FLEX {{source_yyyymm}} Promo Accrual"
    template_cells:
      E26: 'lit("0314")'
      F26: 'lit("000050")'
      H26: 'round(accrual_amount, 2)'
      J26: 'name'
      L26: 'round(due_amount_total, 2)'
      M26: 'percentage'
      E31: 'lit("0314")'
      F31: 'lit("000050")'
      I31: 'round(accrual_amount, 2)'
      J31: 'name'
      L31: 'round(due_amount_total, 2)'
      M31: 'percentage'
  - name: amazon_flex_gray_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0314"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000051"'
    sum_field: "Due Amount"
    percentage: 0.145
    name_template: "Amazon FLEX Gray {{source_yyyymm}} Promo Accrual"
    template_cells:
      E27: 'lit("0314")'
      F27: 'lit("000051")'
      H27: 'round(accrual_amount, 2)'
      J27: 'name'
      L27: 'round(due_amount_total, 2)'
      M27: 'percentage'
      E32: 'lit("0314")'
      F32: 'lit("000051")'
      I32: 'round(accrual_amount, 2)'
      J32: 'name'
      L32: 'round(due_amount_total, 2)'
      M32: 'percentage'
  - name: amazon_skil_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0314"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000070"'
    sum_field: "Due Amount"
    percentage: 0.12
    name_template: "Amazon SKIL {{source_yyyymm}} Promo Accrual"
    template_cells:
      E28: 'lit("0314")'
      F28: 'lit("000070")'
      H28: 'round(accrual_amount, 2)'
      J28: 'name'
      L28: 'round(due_amount_total, 2)'
      M28: 'percentage'
      E33: 'lit("0314")'
      F33: 'lit("000070")'
      I33: 'round(accrual_amount, 2)'
      J33: 'name'
      L33: 'round(due_amount_total, 2)'
      M33: 'percentage'
  - name: amazon_skilsaw_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0314"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000080"'
    sum_field: "Due Amount"
    percentage: 0.12
    name_template: "Amazon SKILSAW {{source_yyyymm}} Promo Accrual"
    template_cells:
      E29: 'lit("0314")'
      F29: 'lit("000080")'
      H29: 'round(accrual_amount, 2)'
      J29: 'name'
      L29: 'round(due_amount_total, 2)'
      M29: 'percentage'
      E34: 'lit("0314")'
      F34: 'lit("000080")'
      I34: 'round(accrual_amount, 2)'
      J34: 'name'
      L34: 'round(due_amount_total, 2)'
      M34: 'percentage'
  - name: lowes_ego_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000030"'
    sum_field: "Due Amount"
    percentage: 0.10
    name_template: "Lowes EGO {{source_yyyymm}} Promo Accrual"
    template_cells:
      E55: 'lit("0178")'
      F55: 'lit("000030")'
      H55: 'round(accrual_amount, 2)'
      J55: 'name'
      L55: 'round(due_amount_total, 2)'
      M55: 'percentage'
      E61: 'lit("0178")'
      F61: 'lit("000030")'
      J61: 'name'
      M61: 'percentage'
  - name: lowes_flex_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000050"'
    sum_field: "Due Amount"
    percentage: 0.08
    name_template: "Lowes FLEX {{source_yyyymm}} Promo Accrual"
    template_cells:
      E56: 'lit("0178")'
      F56: 'lit("000050")'
      H56: 'round(accrual_amount, 2)'
      J56: 'name'
      L56: 'round(due_amount_total, 2)'
      M56: 'percentage'
  - name: lowes_flex_gray_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") == 20'
      - 'text("Brand") == "000051"'
    sum_field: "Due Amount"
    percentage: 0.08
    name_template: "Lowes FLEX Gray {{source_yyyymm}} Promo Accrual"
    template_cells:
      E57: 'lit("0178")'
      F57: 'lit("000051")'
      H57: 'round(accrual_amount, 2)'
      J57: 'name'
      L57: 'round(due_amount_total, 2)'
      M57: 'percentage'
  - name: lowes_skil_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000070"'
    sum_field: "Due Amount"
    percentage: 0.10
    name_template: "Lowes SKIL {{source_yyyymm}} Promo Accrual"
    template_cells:
      E58: 'lit("0178")'
      F58: 'lit("000070")'
      H58: 'round(accrual_amount, 2)'
      J58: 'name'
      L58: 'round(due_amount_total, 2)'
      M58: 'percentage'
  - name: lowes_skilsaw_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "000080"'
    sum_field: "Due Amount"
    percentage: 0.10
    name_template: "Lowes SKILSAW {{source_yyyymm}} Promo Accrual"
    template_cells:
      E59: 'lit("0178")'
      F59: 'lit("000080")'
      H59: 'round(accrual_amount, 2)'
      J59: 'name'
      L59: 'round(due_amount_total, 2)'
      M59: 'percentage'
  - name: lowes_odm_promo_accrual_1
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") == "ODM002"'
    sum_field: "Due Amount"
    percentage: 0.195
    name_template: "Lowes ODM {{source_yyyymm}} Promo Accrual"
    template_cells:
      E60: 'lit("0178")'
      F60: 'lit("ODM002")'
      H60: 'round(accrual_amount, 2)'
      J60: 'name'
      L60: 'round(due_amount_total, 2)'
      M60: 'percentage'
  - name: lowes_promo_due_total_row61
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
      - 'text("Brand") in ("000030", "000050", "000051", "000070", "000080", "ODM002")'
    group_by:
      - Brand
    group_percentages:
      "000030": 0.10
      "000050": 0.08
      "000051": 0.08
      "000070": 0.10
      "000080": 0.10
      "ODM002": 0.195
    sum_field: "Due Amount"
    percentage: 0.10
    name_template: "Lowes EGO {{source_yyyymm}} Promo Accrual"
    summary_template_cells:
      I61: 'round(total_accrual_amount, 2)'
      L61: 'round(total_due_amount_total, 2)'
  - name: lowes_mst_accrual_by_brand
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
    group_by:
      - Brand
    group_row_pairs:
      - [68, 68]
      - [69, 69]
      - [70, 70]
      - [71, 71]
      - [72, 72]
      - [73, 73]
    group_percentages:
      "000030": 0.10
      "000050": 0.08
      "000051": 0.08
      "000070": 0.10
      "000080": 0.10
      "ODM002": 0.195
    sum_field: "Due Amount"
    percentage: 0.10
    name_template: "Lowes {{source_yyyymm}} MST Accrual"
    template_cells:
      E{{row}}: 'lit("0178")'
      F{{row}}: 'brand'
      I{{row}}: 'round(accrual_amount, 2)'
      J{{row}}: 'name'
      L{{row}}: 'round(due_amount_total, 2)'
      M{{row}}: 'percentage'
    summary_template_cells:
      I66: 'round(total_accrual_amount, 2)'
      J66: 'name'
      L66: 'round(total_due_amount_total, 2)'
      H67: 'round(total_accrual_amount, 2)'
      J67: 'name'
      L67: 'round(total_due_amount_total, 2)'
      H74: 'round(total_accrual_amount, 2)'
      J74: 'name'
      L74: 'round(total_due_amount_total, 2)'
  - name: lowes_coop_marketing_accrual_by_brand
    source_filters:
      - 'text("Customer number") == "0178"'
      - 'num("Due Amount") >= 0'
    group_by:
      - Brand
    group_row_pairs:
      - [39, 45]
      - [40, 46]
      - [41, 47]
      - [42, 48]
      - [43, 49]
      - [44, 50]
    sum_field: "Due Amount"
    percentage: 0.0015
    name_template: "Lowes {{source_yyyymm}} Co-op MarketingAccrual{{brand}}"
    template_cells:
      E{{row}}: 'lit("0178")'
      F{{row}}: 'brand'
      H{{row}}: 'round(accrual_amount, 2)'
      J{{row}}: 'name'
      L{{row}}: 'round(due_amount_total, 2)'
      M{{row}}: 'percentage'
      E{{credit_row}}: 'lit("0178")'
      F{{credit_row}}: 'brand'
      I{{credit_row}}: 'round(accrual_amount, 2)'
      J{{credit_row}}: 'name'
      L{{credit_row}}: 'round(due_amount_total, 2)'
      M{{credit_row}}: 'percentage'
""",
                encoding="utf-8",
            )
            config = load_config(config_path)
            result = process_file(input_path=input_path, config=config, output_path=output_path)

            self.assertEqual(result, output_path)
            output_wb = load_workbook(output_path)
            output_ws = output_wb.active
            self.assertEqual(output_ws["E5"].value, "0376")
            self.assertEqual(output_ws["F5"].value, "000030")
            self.assertEqual(float(output_ws["L5"].value), 70.0)
            self.assertAlmostEqual(float(output_ws["H5"].value), 7.7, places=6)
            self.assertEqual(output_ws["J5"].value, "ACE EGO 202605 Promo Accrual")
            self.assertEqual(output_ws["I6"].value, output_ws["H5"].value)
            self.assertEqual(output_ws["J6"].value, output_ws["J5"].value)
            self.assertEqual(output_ws["L6"].value, output_ws["L5"].value)
            self.assertEqual(output_ws["E14"].value, "0376")
            self.assertEqual(output_ws["F14"].value, "000030")
            self.assertAlmostEqual(float(output_ws["H11"].value), 0.7, places=6)
            self.assertAlmostEqual(float(output_ws["I12"].value), 0.7, places=6)
            self.assertAlmostEqual(float(output_ws["H13"].value), 0.7, places=6)
            self.assertAlmostEqual(float(output_ws["I14"].value), 0.7, places=6)
            self.assertEqual(output_ws["J11"].value, "ACEEGO 202605 1% Marketing Allowance Accrual")
            self.assertEqual(output_ws["J12"].value, output_ws["J11"].value)
            self.assertEqual(output_ws["J13"].value, output_ws["J11"].value)
            self.assertEqual(output_ws["J14"].value, output_ws["J11"].value)
            self.assertEqual(float(output_ws["L11"].value), 70.0)
            self.assertEqual(output_ws["M11"].value, 0.01)
            self.assertEqual(output_ws["E19"].value, "B11946")
            self.assertEqual(output_ws["F19"].value, "000030")
            self.assertAlmostEqual(float(output_ws["H19"].value), 5.6, places=6)
            self.assertAlmostEqual(float(output_ws["I20"].value), 5.6, places=6)
            self.assertEqual(output_ws["J19"].value, "G10 EG0 202605 Promo Accrual")
            self.assertEqual(output_ws["J20"].value, output_ws["J19"].value)
            self.assertEqual(float(output_ws["L19"].value), 40.0)
            self.assertEqual(output_ws["M19"].value, 0.14)
            self.assertEqual(output_ws["E25"].value, "0314")
            self.assertEqual(output_ws["F25"].value, "000030")
            self.assertAlmostEqual(float(output_ws["H25"].value), 5.6, places=6)
            self.assertEqual(output_ws["J25"].value, "Amazon EG0 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L25"].value), 40.0)
            self.assertEqual(output_ws["M25"].value, 0.14)
            self.assertEqual(output_ws["E30"].value, "0314")
            self.assertEqual(output_ws["F30"].value, "000030")
            self.assertAlmostEqual(float(output_ws["I30"].value), 5.6, places=6)
            self.assertEqual(output_ws["J30"].value, output_ws["J25"].value)
            self.assertEqual(float(output_ws["L30"].value), 40.0)
            self.assertEqual(output_ws["M30"].value, 0.14)
            self.assertEqual(output_ws["E26"].value, "0314")
            self.assertEqual(output_ws["F26"].value, "000050")
            self.assertAlmostEqual(float(output_ws["H26"].value), 5.8, places=6)
            self.assertEqual(output_ws["J26"].value, "Amazon FLEX 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L26"].value), 40.0)
            self.assertEqual(output_ws["M26"].value, 0.145)
            self.assertEqual(output_ws["E31"].value, "0314")
            self.assertEqual(output_ws["F31"].value, "000050")
            self.assertAlmostEqual(float(output_ws["I31"].value), 5.8, places=6)
            self.assertEqual(output_ws["J31"].value, output_ws["J26"].value)
            self.assertEqual(float(output_ws["L31"].value), 40.0)
            self.assertEqual(output_ws["M31"].value, 0.145)
            self.assertEqual(output_ws["E27"].value, "0314")
            self.assertEqual(output_ws["F27"].value, "000051")
            self.assertAlmostEqual(float(output_ws["H27"].value), 5.8, places=6)
            self.assertEqual(output_ws["J27"].value, "Amazon FLEX Gray 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L27"].value), 40.0)
            self.assertEqual(output_ws["M27"].value, 0.145)
            self.assertEqual(output_ws["E32"].value, "0314")
            self.assertEqual(output_ws["F32"].value, "000051")
            self.assertAlmostEqual(float(output_ws["I32"].value), 5.8, places=6)
            self.assertEqual(output_ws["J32"].value, output_ws["J27"].value)
            self.assertEqual(float(output_ws["L32"].value), 40.0)
            self.assertEqual(output_ws["M32"].value, 0.145)
            self.assertEqual(output_ws["E28"].value, "0314")
            self.assertEqual(output_ws["F28"].value, "000070")
            self.assertAlmostEqual(float(output_ws["H28"].value), 4.8, places=6)
            self.assertEqual(output_ws["J28"].value, "Amazon SKIL 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L28"].value), 40.0)
            self.assertEqual(output_ws["M28"].value, 0.12)
            self.assertEqual(output_ws["E33"].value, "0314")
            self.assertEqual(output_ws["F33"].value, "000070")
            self.assertAlmostEqual(float(output_ws["I33"].value), 4.8, places=6)
            self.assertEqual(output_ws["J33"].value, output_ws["J28"].value)
            self.assertEqual(float(output_ws["L33"].value), 40.0)
            self.assertEqual(output_ws["M33"].value, 0.12)
            self.assertEqual(output_ws["E29"].value, "0314")
            self.assertEqual(output_ws["F29"].value, "000080")
            self.assertAlmostEqual(float(output_ws["H29"].value), 4.8, places=6)
            self.assertEqual(output_ws["J29"].value, "Amazon SKILSAW 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L29"].value), 40.0)
            self.assertEqual(output_ws["M29"].value, 0.12)
            self.assertEqual(output_ws["E34"].value, "0314")
            self.assertEqual(output_ws["F34"].value, "000080")
            self.assertAlmostEqual(float(output_ws["I34"].value), 4.8, places=6)
            self.assertEqual(output_ws["J34"].value, output_ws["J29"].value)
            self.assertEqual(float(output_ws["L34"].value), 40.0)
            self.assertEqual(output_ws["M34"].value, 0.12)
            self.assertEqual(output_ws["E55"].value, "0178")
            self.assertEqual(output_ws["F55"].value, "000030")
            self.assertAlmostEqual(float(output_ws["H55"].value), 2.0, places=6)
            self.assertEqual(output_ws["J55"].value, "Lowes EGO 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L55"].value), 20.0)
            self.assertEqual(output_ws["M55"].value, 0.10)
            self.assertEqual(output_ws["E61"].value, "0178")
            self.assertEqual(output_ws["F61"].value, "000030")
            self.assertAlmostEqual(float(output_ws["I61"].value), 13.1, places=6)
            self.assertEqual(output_ws["J61"].value, output_ws["J55"].value)
            self.assertEqual(float(output_ws["L61"].value), 120.0)
            self.assertEqual(output_ws["M61"].value, 0.10)
            self.assertEqual(output_ws["E56"].value, "0178")
            self.assertEqual(output_ws["F56"].value, "000050")
            self.assertAlmostEqual(float(output_ws["H56"].value), 1.6, places=6)
            self.assertEqual(output_ws["J56"].value, "Lowes FLEX 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L56"].value), 20.0)
            self.assertEqual(output_ws["M56"].value, 0.08)
            self.assertEqual(output_ws["E57"].value, "0178")
            self.assertEqual(output_ws["F57"].value, "000051")
            self.assertAlmostEqual(float(output_ws["H57"].value), 1.6, places=6)
            self.assertEqual(output_ws["J57"].value, "Lowes FLEX Gray 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L57"].value), 20.0)
            self.assertEqual(output_ws["M57"].value, 0.08)
            self.assertEqual(output_ws["E58"].value, "0178")
            self.assertEqual(output_ws["F58"].value, "000070")
            self.assertAlmostEqual(float(output_ws["H58"].value), 2.0, places=6)
            self.assertEqual(output_ws["J58"].value, "Lowes SKIL 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L58"].value), 20.0)
            self.assertEqual(output_ws["M58"].value, 0.10)
            self.assertEqual(output_ws["E59"].value, "0178")
            self.assertEqual(output_ws["F59"].value, "000080")
            self.assertAlmostEqual(float(output_ws["H59"].value), 2.0, places=6)
            self.assertEqual(output_ws["J59"].value, "Lowes SKILSAW 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L59"].value), 20.0)
            self.assertEqual(output_ws["M59"].value, 0.10)
            self.assertEqual(output_ws["E60"].value, "0178")
            self.assertEqual(output_ws["F60"].value, "ODM002")
            self.assertAlmostEqual(float(output_ws["H60"].value), 3.9, places=6)
            self.assertEqual(output_ws["J60"].value, "Lowes ODM 202605 Promo Accrual")
            self.assertEqual(float(output_ws["L60"].value), 20.0)
            self.assertEqual(output_ws["M60"].value, 0.195)
            expected_lowes_mst = [
                ("000030", 68, 0.10, 2.0),
                ("000050", 69, 0.08, 1.6),
                ("000051", 70, 0.08, 1.6),
                ("000070", 71, 0.10, 2.0),
                ("000080", 72, 0.10, 2.0),
                ("ODM002", 73, 0.195, 3.9),
            ]
            for brand, row, percentage, accrual in expected_lowes_mst:
                self.assertEqual(output_ws[f"E{row}"].value, "0178")
                self.assertEqual(output_ws[f"F{row}"].value, brand)
                self.assertAlmostEqual(float(output_ws[f"I{row}"].value), accrual, places=6)
                self.assertEqual(output_ws[f"J{row}"].value, "Lowes 202605 MST Accrual")
                self.assertEqual(float(output_ws[f"L{row}"].value), 20.0)
                self.assertEqual(output_ws[f"M{row}"].value, percentage)
            self.assertAlmostEqual(float(output_ws["I66"].value), 13.1, places=6)
            self.assertEqual(output_ws["J66"].value, "Lowes 202605 MST Accrual")
            self.assertEqual(float(output_ws["L66"].value), 120.0)
            self.assertAlmostEqual(float(output_ws["H67"].value), 13.1, places=6)
            self.assertEqual(output_ws["J67"].value, "Lowes 202605 MST Accrual")
            self.assertEqual(float(output_ws["L67"].value), 120.0)
            self.assertAlmostEqual(float(output_ws["H74"].value), 13.1, places=6)
            self.assertEqual(output_ws["J74"].value, "Lowes 202605 MST Accrual")
            self.assertEqual(float(output_ws["L74"].value), 120.0)
            expected_lowes = [
                ("000030", 39, 45),
                ("000050", 40, 46),
                ("000051", 41, 47),
                ("000070", 42, 48),
                ("000080", 43, 49),
                ("ODM002", 44, 50),
            ]
            for brand, dr_row, cr_row in expected_lowes:
                self.assertEqual(output_ws[f"E{dr_row}"].value, "0178")
                self.assertEqual(output_ws[f"F{dr_row}"].value, brand)
                self.assertAlmostEqual(float(output_ws[f"H{dr_row}"].value), 0.03, places=6)
                self.assertEqual(output_ws[f"J{dr_row}"].value, f"Lowes 202605 Co-op MarketingAccrual{brand}")
                self.assertEqual(float(output_ws[f"L{dr_row}"].value), 20.0)
                self.assertEqual(output_ws[f"M{dr_row}"].value, 0.0015)
                self.assertEqual(output_ws[f"E{cr_row}"].value, "0178")
                self.assertEqual(output_ws[f"F{cr_row}"].value, brand)
                self.assertAlmostEqual(float(output_ws[f"I{cr_row}"].value), 0.03, places=6)
                self.assertEqual(output_ws[f"J{cr_row}"].value, f"Lowes 202605 Co-op MarketingAccrual{brand}")
                self.assertEqual(float(output_ws[f"L{cr_row}"].value), 20.0)
                self.assertEqual(output_ws[f"M{cr_row}"].value, 0.0015)
            self.assertIn("A4:G4", {str(rng) for rng in output_ws.merged_cells.ranges})


if __name__ == "__main__":
    unittest.main()
